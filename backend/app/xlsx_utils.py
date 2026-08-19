"""Helper Excel (.xlsx) — build & parse workbook via openpyxl.

Dipake kabeh export/import supaya file spreadsheet kebukak rapi ing Excel,
ora CSV maneh.
"""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MIME = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")

# Header biru entheng biar template/export katon jelas
_HEADER_FILL = PatternFill("solid", fgColor="DCE9F7")
_HEADER_FONT = Font(bold=True)


def rows_to_xlsx(headers: list[str], rows: list[list],
                 sheet_title: str = "Data", autosize: bool = True) -> bytes:
    """Gawe file .xlsx saka list of rows (nilai otomatis dadi str)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Data"

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for row in rows:
        ws.append(["" if v is None else str(v) for v in row])

    if autosize:
        for c in range(1, len(headers) + 1):
            letter = get_column_letter(c)
            width = max((len(str(cell.value or "")) for cell in ws[letter]),
                        default=8)
            ws.column_dimensions[letter].width = min(width + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_to_rows(content: bytes) -> tuple[list[str], list[list]]:
    """Parse .xlsx → (headers, rows). Ngelola data_only + nilai str kosong."""
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = [[_clean(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    if not rows:
        return [], []
    headers = [str(h or "").strip() for h in rows[0]]
    return headers, rows[1:]


def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()
