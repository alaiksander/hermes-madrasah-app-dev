"""Murid view: list + form + hapus + bulk action + import/export Excel."""
import io
import httpx
import openpyxl

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import RedirectResponse, Response

from ....core.client import (api_delete, api_get, api_get_raw, api_patch,
                             api_post, api_post_multipart)
from ....core.deps import get_current_user_web, require_login_web, require_permission_web
from ....core.templates import templates

router = APIRouter(tags=["web-data-murid"])


def _redirect(msg: str, type_: str = "success", path: str = "/madrasah-app/data/murid"):
    return RedirectResponse(
        url=f"{path}?msg={msg.replace(' ', '+')}&type={type_}",
        status_code=303,
    )


@router.get("")
async def murid_list(
    request: Request,
    q: str | None = None,
    kelas_id: int | None = None,
    tahun_ajaran_id: int | None = None,
    page: int = 1,
    per_page: int = 50,
    semua: bool = False,
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import", "murid.qr")),
):
    """Daftar murid dengan search, filter kelas + tahun ajaran, pagination.

    Tahun ajaran: default ke yang aktif. Bisa override via query param.
    Kalau request dari HTMX (header HX-Request: true), return partial
    saja — supaya swap #murid-table-wrap tanpa duplikat filter bar.
    """
    # Fetch tahun ajaran untuk dropdown + auto-filter
    ta_r = await api_get(request, "/api/tahun-ajaran")
    tahun_ajaran_list = ta_r.json() if ta_r.status_code == 200 else []
    tahun_aktif = next((ta for ta in tahun_ajaran_list if ta.get("is_active")), None)
    if tahun_ajaran_id is None and tahun_aktif:
        tahun_ajaran_id = tahun_aktif["id"]

    # Filter kelas sesuai tahun ajaran (default aktif) — supaya dropdown
    # tidak duplikat nama_kelas antar tahun ajaran.
    kelas_r = await api_get(request, "/api/kelas", tahun_ajaran_id=tahun_ajaran_id or "")
    kelas_all_r = await api_get(request, "/api/kelas")  # semua kelas (semua TA)
    kelas_list = kelas_r.json() if kelas_r.status_code == 200 else []
    kelas_all = kelas_all_r.json() if kelas_all_r.status_code == 200 else []

    # Fetch murid dengan filter
    params = {"page": page, "per_page": per_page}
    if q:
        params["q"] = q
    if kelas_id:
        params["kelas_id"] = kelas_id

    murid_r = await api_get(request, "/api/murid", **params)

    if murid_r.status_code != 200:
        return _redirect(f"Gagal memuat murid ({murid_r.status_code})", "error")

    data = murid_r.json()

    ctx = {
        "user": user,
        "murid": data.get("items", []),
        "total": data.get("total", 0),
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, -(-data.get("total", 0) // per_page)),
        "q": q or "",
        "kelas_id": kelas_id,
        "kelas_list": kelas_list,
        "kelas_tahun_map": {str(k["id"]): str(k.get("tahun_ajaran_id") or "")
                           for k in kelas_all if k.get("id") is not None},
        "tahun_ajaran_list": tahun_ajaran_list,
        "tahun_aktif_id": tahun_aktif["id"] if tahun_aktif else None,
        "tahun_aktif_nama": tahun_aktif["nama"] if tahun_aktif else "",
        "semua": semua,
    }

    if request.headers.get("HX-Request") == "true":
        return templates.TemplateResponse(request, "murid/_table.html", ctx)

    return templates.TemplateResponse(request, "murid/list.html", ctx)


@router.get("/baru")
async def murid_baru(
    request: Request,
    user: dict = Depends(require_permission_web("murid.view")),
):
    """Form tambah murid baru (admin only)."""
    kelas_r = await api_get(request, "/api/kelas")
    kelas_list = kelas_r.json() if kelas_r.status_code == 200 else []
    return templates.TemplateResponse(
        request,
        "murid/form.html",
        {
            "user": user,
            "murid": None,
            "kelas_list": kelas_list,
            "form_title": "Tambah Murid",
            "form_action": "/madrasah-app/data/murid",
        },
    )


@router.get("/{mid:int}")
async def murid_edit(
    request: Request,
    mid: int,
    user: dict = Depends(require_permission_web("murid.view")),
):
    """Form edit murid (admin only)."""
    r = await api_get(request, f"/api/murid/{mid}")
    if r.status_code != 200:
        return _redirect("Murid tidak ditemukan", "error")
    kelas_r = await api_get(request, "/api/kelas")
    kelas_list = kelas_r.json() if kelas_r.status_code == 200 else []
    return templates.TemplateResponse(
        request,
        "murid/form.html",
        {
            "user": user,
            "murid": r.json(),
            "kelas_list": kelas_list,
            "form_title": "Edit Murid",
            "form_action": f"/madrasah-app/data/murid/{mid}",
        },
    )


@router.post("")
async def murid_create(
    request: Request,
    nisn: str = Form(""),
    nama: str = Form(...),
    kelas_id: int = Form(...),
    nama_ortu: str = Form(""),
    telepon: str = Form(""),
    nik: str = Form(""),
    tempat_lahir: str = Form(""),
    tanggal_lahir: str = Form(""),
    jenis_kelamin: str = Form(""),
    alamat: str = Form(""),
    nama_ayah_kandung: str = Form(""),
    nama_ibu_kandung: str = Form(""),
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import")),
):
    """Submit form tambah murid → POST API existing."""
    payload = {
        "nisn": nisn.strip() or None,
        "nama": nama.strip(),
        "kelas_id": kelas_id,
        "nama_ortu": nama_ortu.strip() or None,
        "telepon": telepon.strip() or None,
        "nik": nik.strip() or None,
        "tempat_lahir": tempat_lahir.strip() or None,
        "tanggal_lahir": tanggal_lahir.strip() or None,
        "jenis_kelamin": jenis_kelamin.strip() or None,
        "alamat": alamat.strip() or None,
        "nama_ayah_kandung": nama_ayah_kandung.strip() or None,
        "nama_ibu_kandung": nama_ibu_kandung.strip() or None,
    }
    r = await api_post(request, "/api/murid", json=payload)
    if r.status_code == 201:
        return _redirect(f"Murid {nama} berhasil ditambahkan")
    # Handle error — re-render form dengan pesan
    detail = "Gagal menambahkan murid"
    try:
        detail = r.json().get("detail", detail)
    except Exception:
        pass
    kelas_r = await api_get(request, "/api/kelas")
    return templates.TemplateResponse(
        request,
        "murid/form.html",
        {
            "user": user,
            "murid": {"nisn": nisn, "nama": nama, "kelas_id": kelas_id,
                      "nama_ortu": nama_ortu, "telepon": telepon,
                      "nik": nik, "tempat_lahir": tempat_lahir,
                      "tanggal_lahir": tanggal_lahir,
                      "jenis_kelamin": jenis_kelamin, "alamat": alamat,
                      "nama_ayah_kandung": nama_ayah_kandung,
                      "nama_ibu_kandung": nama_ibu_kandung},
            "kelas_list": kelas_r.json() if kelas_r.status_code == 200 else [],
            "form_title": "Tambah Murid",
            "form_action": "/madrasah-app/data/murid",
            "error": detail,
        },
        status_code=400,
    )


@router.post("/{mid:int}")
async def murid_update(
    request: Request,
    mid: int,
    nisn: str = Form(""),
    nama: str = Form(...),
    kelas_id: int = Form(...),
    nama_ortu: str = Form(""),
    telepon: str = Form(""),
    nik: str = Form(""),
    tempat_lahir: str = Form(""),
    tanggal_lahir: str = Form(""),
    jenis_kelamin: str = Form(""),
    alamat: str = Form(""),
    nama_ayah_kandung: str = Form(""),
    nama_ibu_kandung: str = Form(""),
    is_active: str = Form(""),
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import")),
):
    """Submit form edit murid → PATCH API existing."""
    payload = {
        "nisn": nisn.strip() or None,
        "nama": nama.strip(),
        "kelas_id": kelas_id,
        "nama_ortu": nama_ortu.strip() or None,
        "telepon": telepon.strip() or None,
        "nik": nik.strip() or None,
        "tempat_lahir": tempat_lahir.strip() or None,
        "tanggal_lahir": tanggal_lahir.strip() or None,
        "jenis_kelamin": jenis_kelamin.strip() or None,
        "alamat": alamat.strip() or None,
        "nama_ayah_kandung": nama_ayah_kandung.strip() or None,
        "nama_ibu_kandung": nama_ibu_kandung.strip() or None,
    }
    if is_active in ("true", "false"):
        payload["is_active"] = is_active == "true"

    r = await api_patch(request, f"/api/murid/{mid}", json=payload)
    if r.status_code == 200:
        return _redirect(f"Murid {nama} berhasil diperbarui")
    detail = "Gagal memperbarui murid"
    try:
        detail = r.json().get("detail", detail)
    except Exception:
        pass
    return _redirect(detail, "error")


@router.post("/murid-bulk-archive")
async def murid_bulk_archive(
    request: Request,
    ids: str = Form(...),
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import")),
):
    """Arsipkan banyak murid sekaligus. `ids` = comma-separated.

    PENTING: pakai path `/murid-bulk-archive` (dengan dash, BUKAN slash
    `/murid/bulk-archive`) — kalau pakai slash, FastAPI match ke route
    parameterized `/murid/{mid}` lebih dulu karena prioritas regex prefix.
    """
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return _redirect("Tidak ada murid yang dipilih", "error")
    success = 0
    for mid in id_list:
        r = await api_delete(request, f"/api/murid/{mid}")
        if r.status_code == 200:
            success += 1
    return _redirect(f"{success} dari {len(id_list)} murid berhasil diarsipkan")


@router.post("/{mid:int}/hapus")
async def murid_hapus(
    request: Request,
    mid: int,
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import")),
):
    """Soft-delete (archive) murid."""
    r = await api_delete(request, f"/api/murid/{mid}")
    if r.status_code == 200:
        return _redirect("Murid berhasil diarsipkan")
    detail = "Gagal mengarsipkan murid"
    try:
        detail = r.json().get("detail", detail)
    except Exception:
        pass
    return _redirect(detail, "error")


@router.post("/{mid:int}/aktifkan")
async def murid_aktifkan(
    request: Request,
    mid: int,
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import")),
):
    """Re-aktifkan murid yang diarsipkan."""
    r = await api_patch(request, f"/api/murid/{mid}", json={"is_active": True})
    if r.status_code == 200:
        return _redirect("Murid berhasil diaktifkan kembali")
    detail = "Gagal mengaktifkan murid"
    try:
        detail = r.json().get("detail", detail)
    except Exception:
        pass
    return _redirect(detail, "error")





@router.get("/{mid:int}/detail")
async def murid_detail(
    request: Request,
    mid: int,
    user: dict = Depends(require_login_web),
):
    """Halaman detail murid (read-only) — termasuk section BK ringkas.

    Akses: login (murid.view). Gunakan untuk admin cek profil + melihat
    summary BK tanpa masuk ke halaman edit. Link ke monitor BK lengkap.
    """
    r = await api_get(request, f"/api/murid/{mid}")
    if r.status_code != 200:
        return _redirect("Murid tidak ditemukan", "error")
    murid = r.json()

    # Format tanggal lahir: ISO "2014-03-30" → "30/03/2014" untuk tampilan
    tgl_lahir = murid.get("tanggal_lahir")
    if tgl_lahir:
        try:
            from datetime import datetime as _dt
            murid["tanggal_lahir"] = _dt.strptime(str(tgl_lahir)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            pass

    # Section BK (summary) — hanya kalau user punya bk.view
    bk_summary = None
    if user.get("role") == "admin" or user.get("role_id") is not None or        any(k in str(user) for k in ["bk"]):
        # Backend sudah kasih role_id di user dict (Fase-2)
        pass
    # Selalu coba ambil BK summary (admin selalu boleh, guru tergantung permission)
    try:
        bk_r = await api_get(request, f"/api/bk/monitor/{mid}")
        if bk_r.status_code == 200:
            data = bk_r.json()
            bk_summary = {
                "total_catatan": data["rekap"]["total_catatan"],
                "total_poin_pelanggaran": data["rekap"]["total_poin_pelanggaran"],
                "total_poin_prestasi": data["rekap"]["total_poin_prestasi"],
                "status_sp": data["rekap"]["status_sp"],
                "total_sesi": data["rekap"]["total_sesi"],
                "catatan_terbaru": data.get("catatan", [])[:5],
                "sesi_terbaru": data.get("sesi", [])[:5],
            }
    except Exception:
        pass

    return templates.TemplateResponse(
        request,
        "murid/detail.html",
        {
            "user": user,
            "murid": murid,
            "bk_summary": bk_summary,
        },
    )


# ── Import / Export

# ── Import / Export Excel (Task 1.6) ──────────────────────────────────


@router.get("/template-xlsx")
async def murid_template(
    request: Request,
    user: dict = Depends(require_permission_web("murid.view")),
):
    """Download template Excel import (admin only).

    PENTING: pakai dash `/murid-template-xlsx` (BUKAN slash + dot
    `/murid/template.xlsx`) — kalau pakai slash+dot, FastAPI match
    ke route parameterized `/murid/{mid}` dan gagal int parse.
    """
    content = await api_get_raw(request, "/api/murid/template.xlsx")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template-import-murid.xlsx"'},
    )


@router.get("/export-xlsx")
async def murid_export(
    request: Request,
    user: dict = Depends(require_permission_web("murid.view")),
):
    """Export SEMUA murid aktif ke Excel (admin only). Path pakai dash
    untuk hindari conflict dengan route /murid/{mid}."""
    content = await api_get_raw(request, "/api/murid/export.xlsx")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="daftar-murid.xlsx"'},
    )


@router.get("/import")
async def murid_import_page(
    request: Request,
    user: dict = Depends(require_permission_web("murid.view")),
):
    """Halaman upload Excel import (admin only).

    PENTING: pakai dash `/murid-import` (BUKAN slash `/murid/import`)
    untuk hindari conflict dengan route `/murid/{mid}`.
    """
    kelas_r = await api_get(request, "/api/kelas")
    return templates.TemplateResponse(
        request,
        "murid/import.html",
        {
            "user": user,
            "kelas_list": kelas_r.json() if kelas_r.status_code == 200 else [],
            "preview": None,
        },
    )


@router.post("/import-preview")
async def murid_import_preview(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import")),
):
    """Upload Excel → parse → tampilkan preview (belum commit ke DB).

    Pakai openpyxl langsung (tidak commit ke DB) untuk preview cepat.
    PENTING: pakai dash `/murid-import-preview` untuk hindari conflict.

    Catatan: hanya row dengan kelas yang SUDAH ADA di database yang lolos validasi,
    karena API existing /api/murid/import auto-create kelas tanpa tahun_ajaran_id
    (IntegrityError kalau tahun_ajaran_id null).
    """
    kelas_r = await api_get(request, "/api/kelas")
    kelas_list = kelas_r.json() if kelas_r.status_code == 200 else []

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        return templates.TemplateResponse(
            request,
            "murid/import.html",
            {
                "user": user,
                "kelas_list": kelas_list,
                "preview": None,
                "error": f"File bukan Excel valid: {e}",
                "filename": file.filename,
            },
            status_code=400,
        )

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return templates.TemplateResponse(
            request,
            "murid/import.html",
            {
                "user": user,
                "kelas_list": kelas_list,
                "preview": None,
                "error": "File Excel kosong",
                "filename": file.filename,
            },
            status_code=400,
        )

    headers = [str(h or "").strip().lower() for h in rows[0]]
    norm_headers = [h.replace("_", " ").replace("-", " ").strip() for h in headers]

    header_map = {
        "nisn": "nisn", "nis": "nisn", "nama": "nama", "kelas": "nama_kelas",
        "nama ortu": "nama_ortu", "telepon": "telepon", "wa ortu": "telepon",
    }
    field_names = [header_map.get(h, h) for h in norm_headers]

    # Build set nama kelas yang valid (existing)
    existing_kelas = {k["nama_kelas"] for k in kelas_list}

    data_rows = []
    for r_idx, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        item = {}
        for f_name, cell in zip(field_names, row):
            if f_name:
                item[f_name] = str(cell or "").strip()

        errors = []
        nisn_val = item.get("nisn", "")
        if nisn_val and not (nisn_val.isdigit() and len(nisn_val) == 10):
            errors.append("NISN harus 10 digit angka (boleh kosong)")
        if not item.get("nama"):
            errors.append("Nama kosong")
        if not item.get("nama_kelas"):
            errors.append("Kelas kosong")
        elif item["nama_kelas"] not in existing_kelas:
            errors.append(f'Kelas "{item["nama_kelas"]}" belum ada di sistem (tambah dulu)')
        # Telepon format check (opsional)
        telp = item.get("telepon", "")
        if telp and not (telp.replace("+", "").isdigit() and 8 <= len(telp) <= 15):
            errors.append("Telepon format tidak valid (8-15 digit)")

        item["_row"] = r_idx
        item["_errors"] = errors
        item["_valid"] = len(errors) == 0
        data_rows.append(item)

    valid_count = sum(1 for r in data_rows if r["_valid"])

    return templates.TemplateResponse(
        request,
        "murid/import.html",
        {
            "user": user,
            "kelas_list": kelas_list,
            "preview": {
                "headers": headers,
                "rows": data_rows,
                "filename": file.filename,
                "valid_count": valid_count,
                "total_count": len(data_rows),
            },
        },
    )


@router.post("/import")
async def murid_import_commit(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.update", "murid.delete", "murid.import")),
):
    """Submit Excel → forward ke API existing /api/murid/import → flash message.

    PENTING: pakai dash `/murid-import` untuk hindari conflict dengan `/murid/{mid}`.

    Catatan: API existing /api/murid/import auto-create kelas tanpa tahun_ajaran_id
    (IntegrityError). Jadi di web, FILTER row yang kelas-nya tidak ada di sistem
    sebelum forward — supaya tidak ada error di API existing.
    """
    contents = await file.read()

    # Parse Excel — filter baris invalid (kelas tidak ada)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return _redirect(f"File bukan Excel valid: {e}", "error")

    if not all_rows:
        return _redirect("File Excel kosong", "error")

    # Get existing kelas
    kelas_r = await api_get(request, "/api/kelas")
    kelas_list = kelas_r.json() if kelas_r.status_code == 200 else []
    existing_kelas = {k["nama_kelas"] for k in kelas_list}

    headers = [str(h or "").strip().lower() for h in all_rows[0]]
    norm_headers = [h.replace("_", " ").replace("-", " ").strip() for h in headers]
    header_map = {
        "nisn": "nisn", "nis": "nisn", "nama": "nama", "kelas": "nama_kelas",
        "nama ortu": "nama_ortu", "telepon": "telepon", "wa ortu": "telepon",
    }
    field_names = [header_map.get(h, h) for h in norm_headers]

    # Build cleaned workbook dengan hanya row valid
    cleaned_wb = openpyxl.Workbook()
    cleaned_ws = cleaned_wb.active
    cleaned_ws.append(all_rows[0])  # header

    kept = 0
    for row in all_rows[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        item = {}
        for f_name, cell in zip(field_names, row):
            if f_name:
                item[f_name] = str(cell or "").strip()
        # Filter: skip kalau kelas tidak ada
        if not item.get("nama_kelas") or item["nama_kelas"] not in existing_kelas:
            continue
        # Filter: skip kalau NISN diisi tapi format salah (NISN boleh kosong)
        nisn_val = item.get("nisn", "")
        if nisn_val and not (nisn_val.isdigit() and len(nisn_val) == 10):
            continue
        # Filter: skip kalau nama kosong
        if not item.get("nama"):
            continue
        # Filter: skip kalau telepon format invalid (kalau ada)
        telp = item.get("telepon", "")
        if telp and not (telp.replace("+", "").isdigit() and 8 <= len(telp) <= 15):
            continue
        cleaned_ws.append(row)
        kept += 1

    if kept == 0:
        return _redirect("Tidak ada baris valid untuk diimpor", "error")

    # Save cleaned workbook to bytes
    cleaned_buf = io.BytesIO()
    cleaned_wb.save(cleaned_buf)
    cleaned_bytes = cleaned_buf.getvalue()

    # Forward ke API existing
    import httpx
    async with httpx.AsyncClient(timeout=60) as c:
        token = request.cookies.get("madrasah_app_token")
        r = await c.post(
            "http://127.0.0.1:8010/api/murid/import",
            files={"file": (file.filename or "import.xlsx", cleaned_bytes,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"Authorization": f"Bearer {token}"} if token else {},
        )

    if r.status_code == 200:
        try:
            result = r.json()
            added = result.get("added", result.get("ditambahkan", 0))
            skipped = result.get("skipped", result.get("sudah_ada", 0))
            error_count = len(result.get("error", []))
            msg = f"Import selesai: {added} ditambah"
            if skipped:
                msg += f", {skipped} dilewati (NISN duplikat)"
            if error_count:
                msg += f", {error_count} error"
        except Exception:
            msg = "Import selesai"
        return _redirect(msg)
    detail = "Gagal import"
    try:
        detail = r.json().get("detail", detail)
    except Exception:
        pass
    return _redirect(detail, "error")


# ── QR Card + Cetak Massal A4 (Task 1.7) ──────────────────────────────
# Halaman QR (per kelas & per anak) pindah ke menu Kartu QR (kartu_qr.py,
# 2026-08-15) — di sini hanya proxy PNG/PDF yang dipakai preview.


@router.get("/{mid:int}/qr-png")
async def murid_qr_png(
    request: Request,
    mid: int,
    user: dict = Depends(require_login_web),
):
    """Proxy QR PNG dari API existing."""
    try:
        content = await api_get_raw(request, f"/api/murid/{mid}/qr.png")
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 404:
            return templates.TemplateResponse(
                request,
                "error.html",
                {"status_code": 404, "message": "Murid tidak ditemukan."},
                status_code=404,
            )
        raise
    return Response(
        content=content,
        media_type="image/png",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@router.get("/{mid:int}/qr-pdf")
async def murid_qr_pdf(
    request: Request,
    mid: int,
    user: dict = Depends(require_permission_web("murid.view")),
):
    """PDF QR Card 1 murid (admin only)."""
    try:
        content = await api_get_raw(request, f"/api/murid/{mid}/qr.pdf")
    except httpx.HTTPStatusError as e:
        if e.response.status_code == 404:
            return templates.TemplateResponse(
                request,
                "error.html",
                {"status_code": 404, "message": "Murid tidak ditemukan."},
                status_code=404,
            )
        raise
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="qr-{mid}.pdf"'},
    )


# ─────────────────────────────────────────────────────────────
# Import EMIS Kemenag (nilai jual: ekspor EMIS langsung masuk)
# Route pakai dash untuk hindari conflict /murid/{mid}.
# ─────────────────────────────────────────────────────────────
import re
import uuid as uuid_mod

_EMIS_KEY_ALIASES = {
    "nama": ("nama lengkap", "nama"),
    "nisn": ("nisn",),
    "nik": ("nik", "nik ktp"),
    "tempat_lahir": ("tempat lahir",),
    "tanggal_lahir": ("tanggal lahir", "ttl tanggal"),
    "tingkat_rombel": ("tingkat - rombel", "tingkat rombel", "rombel", "kelas"),
    "jenis_kelamin": ("jenis kelamin", "jk"),
    "alamat": ("alamat",),
    "telepon": ("no telepon", "telepon", "no hp", "no telp"),
    "ayah": ("nama ayah kandung", "nama ayah", "ayah kandung"),
    "ibu": ("nama ibu kandung", "nama ibu", "ibu kandung"),
    "nama_ortu": ("nama wali", "nama ortu", "nama orang tua"),
}


def _parse_rombel_web(raw: str) -> tuple[int | None, str]:
    """Parse 'Kelas 7 - 01' → (7, '01'). Return (None, raw) kalau gagal."""
    m = re.search(r"(\d+)\s*[-–]\s*(\d+)", str(raw or ""))
    if m:
        return int(m.group(1)), m.group(2)
    m2 = re.search(r"(\d+)", str(raw or ""))
    if m2:
        return int(m2.group(1)), ""
    return None, str(raw or "").strip()


def _auto_kelas_web(tingkat: int, rombel: str, kelas_list: list[dict]) -> str:
    """Auto-map rombel → nama kelas existing (pola ABC dulu, lalu angka)."""
    nama_kelas_existing = [k.get("nama_kelas", "") for k in kelas_list]
    # Pola ABC: 01 → A, 02 → B ...
    if rombel.isdigit():
        huruf = chr(ord("A") + int(rombel) - 1)
        for suffix in (f"{tingkat}{huruf}", f"{tingkat} {huruf}"):
            if suffix in nama_kelas_existing:
                return suffix
        # Pola angka: 7-01 / 7.1 / 7-1
        for cand in (f"{tingkat}-{int(rombel):02d}", f"{tingkat}.{int(rombel)}",
                     f"{tingkat}-{int(rombel)}"):
            if cand in nama_kelas_existing:
                return cand
    # Fallback: buat baru pakai format EMIS (7-01)
    if rombel.isdigit():
        return f"{tingkat}-{int(rombel):02d}"
    return f"{tingkat}-{rombel}" if tingkat else "Kelas Baru"


@router.get("/import-emis")
async def murid_import_emis_page(
    request: Request,
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.import")),
):
    """Halaman import murid dari ekspor EMIS Kemenag (nilai jual)."""
    return templates.TemplateResponse(
        request,
        "murid/import_emis.html",
        {"user": user, "preview": None, "error": None, "result": None},
    )


@router.post("/import-emis-preview")
async def murid_import_emis_preview(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.import")),
):
    """Upload file EMIS → parse → preview per sheet + auto-map kelas."""
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        return templates.TemplateResponse(
            request, "murid/import_emis.html",
            {"user": user, "preview": None, "error": f"File bukan Excel EMIS valid: {e}",
             "result": None}, status_code=400)

    kelas_r = await api_get(request, "/api/kelas")
    kelas_list = kelas_r.json() if kelas_r.status_code == 200 else []

    sheets = []
    total_valid = 0
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h or "").strip().lower().replace("_", " ").replace("-", " ")
                   for h in rows[0]]
        idx: dict[str, int] = {}
        for key, aliases in _EMIS_KEY_ALIASES.items():
            for a in aliases:
                if a in headers:
                    idx[key] = headers.index(a)
                    break
        if "nama" not in idx and "nisn" not in idx:
            continue  # bukan sheet murid

        n_valid = 0
        rombel_raw = ""
        for r in rows[1:]:
            if all(c is None or str(c).strip() == "" for c in r):
                continue
            if "nama" in idx and r[idx["nama"]] is not None and str(r[idx["nama"]]).strip():
                n_valid += 1
            elif "nisn" in idx and r[idx["nisn"]] is not None and str(r[idx["nisn"]]).strip():
                n_valid += 1
            if not rombel_raw and "tingkat_rombel" in idx:
                rombel_raw = str(r[idx["tingkat_rombel"]] or "").strip()
        rombel_raw = rombel_raw or ws.title
        tingkat, rombel = _parse_rombel_web(rombel_raw)
        kelas_target = _auto_kelas_web(tingkat, rombel, kelas_list) if tingkat else rombel_raw
        total_valid += n_valid
        sheets.append({
            "nama": ws.title,
            "rombel": rombel_raw,
            "jumlah": n_valid,
            "kelas_target": kelas_target,
            "tingkat": tingkat,
        })

    if not sheets:
        return templates.TemplateResponse(
            request, "murid/import_emis.html",
            {"user": user, "preview": None,
             "error": "Tidak ada sheet berisi data murid EMIS (kolom Nama/NISN tidak ditemukan).",
             "result": None}, status_code=400)

    # Simpan file sementara untuk commit (hapus setelah dipakai)
    tmp_path = f"/tmp/madrasah_emis_{uuid_mod.uuid4().hex}.xlsx"
    with open(tmp_path, "wb") as f:
        f.write(contents)

    return templates.TemplateResponse(
        request, "murid/import_emis.html",
        {"user": user, "preview": {
            "filename": file.filename,
            "sheets": sheets,
            "total_valid": total_valid,
            "tmp_path": tmp_path,
        }, "error": None, "result": None})


@router.post("/import-emis-commit")
async def murid_import_emis_commit(
    request: Request,
    tmp_path: str = Form(""),
    user: dict = Depends(require_permission_web("murid.view", "murid.create", "murid.import")),
):
    """Commit: kirim file EMIS tersimpan ke API /api/murid/import-emis."""
    if not tmp_path or not tmp_path.startswith("/tmp/madrasah_emis_"):
        return templates.TemplateResponse(
            request, "murid/import_emis.html",
            {"user": user, "preview": None,
             "error": "File sementara tidak valid — upload ulang.",
             "result": None}, status_code=400)

    try:
        with open(tmp_path, "rb") as f:
            content = f.read()
    except FileNotFoundError:
        return templates.TemplateResponse(
            request, "murid/import_emis.html",
            {"user": user, "preview": None,
             "error": "File sementara sudah kedaluwarsa — upload ulang.",
             "result": None}, status_code=400)

    try:
        r = await api_post_multipart(
            request, "/api/murid/import-emis",
            files={"file": ("daftar-siswa-emis.xlsx", content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        data = r.json()
    except Exception as e:
        return templates.TemplateResponse(
            request, "murid/import_emis.html",
            {"user": user, "preview": None, "error": f"Gagal import: {e}", "result": None},
            status_code=502)

    import os
    try:
        os.remove(tmp_path)
    except OSError:
        pass

    if r.status_code != 200:
        return templates.TemplateResponse(
            request, "murid/import_emis.html",
            {"user": user, "preview": None,
             "error": data.get("detail", f"Gagal import (HTTP {r.status_code})"),
             "result": None}, status_code=400)

    return templates.TemplateResponse(
        request, "murid/import_emis.html",
        {"user": user, "preview": None, "error": None, "result": data})