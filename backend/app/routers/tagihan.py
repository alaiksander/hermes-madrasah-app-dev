"""Router Tagihan & Pembayaran — modul Pembayaran.

Alur: JenisPembayaran (master) → Tagihan (per murid per bulan) → Pembayaran
(lunas/cicil). Fitur: auto-generate bulanan, keringanan (potongan Rp),
penundaan (geser jatuh tempo), rekap per kelas.
"""
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import func, select
from sqlalchemy.orm import Session, joinedload

from ..deps import get_tenant_db, require_permission
from ..models import (Guru, JenisPembayaran, Kelas, Murid, Pembayaran,
                      Tagihan)
from ..schemas import (JenisPembayaranCreate, JenisPembayaranOut,
                       JenisPembayaranUpdate, TagihanBayarCreate,
                       TagihanCreate, TagihanKeringananCreate, TagihanOut,
                       TagihanTundaCreate)

router = APIRouter(prefix="/api/tagihan", tags=["Tagihan & Pembayaran"])
WIB = ZoneInfo("Asia/Jakarta")


def _tagihan_out(t: Tagihan) -> dict:
    dibayar = sum(p.nominal for p in t.pembayaran)
    sisa = t.nominal - t.potongan - dibayar
    return {
        "id": t.id,
        "murid_id": t.murid_id,
        "jenis_id": t.jenis_id,
        "periode": t.periode,
        "nominal": t.nominal,
        "potongan": t.potongan,
        "jatuh_tempo": t.jatuh_tempo,
        "ditunda_sampai": t.ditunda_sampai,
        "status": t.status,
        "catatan": t.catatan,
        "murid_nama": t.murid.nama if t.murid else "",
        "murid_nisn": t.murid.nisn if t.murid else "",
        "murid_kelas": t.murid.kelas.nama_kelas if t.murid and t.murid.kelas else "",
        "jenis_nama": t.jenis.nama if t.jenis else "",
        "jenis_periode": t.jenis.periode if t.jenis else "",
        "dibayar": dibayar,
        "sisa": max(sisa, 0),
        "pembayaran": [
            {"id": p.id, "tagihan_id": p.tagihan_id, "nominal": p.nominal,
             "metode": p.metode, "tanggal": p.tanggal,
             "guru_id": p.guru_id, "catatan": p.catatan,
             "guru_nama": p.guru.nama if p.guru else ""}
            for p in t.pembayaran
        ],
    }


def _refresh_status(db: Session, t: Tagihan) -> None:
    """Status diturunkan dari pembayaran (via query langsung — anti stale)."""
    dibayar = (db.query(func.coalesce(func.sum(Pembayaran.nominal), 0))
               .filter(Pembayaran.tagihan_id == t.id).scalar()) or 0
    sisa = t.nominal - t.potongan - dibayar
    if t.ditunda_sampai and t.ditunda_sampai.date() > date.today():
        t.status = "ditunda"
    elif sisa <= 0:
        t.status = "lunas"
    elif dibayar > 0:
        t.status = "sebagian"
    else:
        t.status = "belum"


# ─── Master: Jenis Pembayaran ─────────────────────────────────────────────

@router.get("/jenis", response_model=list[JenisPembayaranOut])
def list_jenis(
    user: dict = Depends(require_permission("tagihan.view", "tagihan.kelola")),
    db: Session = Depends(get_tenant_db),
):
    """Daftar semua jenis pembayaran (dengan yang non-aktif)."""
    rows = (db.query(JenisPembayaran)
            .order_by(JenisPembayaran.is_active.desc(), JenisPembayaran.nama)
            .all())
    return rows


@router.post("/jenis", response_model=JenisPembayaranOut, status_code=201)
def create_jenis(
    data: JenisPembayaranCreate,
    user: dict = Depends(require_permission("tagihan.kelola")),
    db: Session = Depends(get_tenant_db),
):
    """Buat jenis pembayaran baru."""
    if data.periode not in ("bulanan", "sekali", "semester"):
        raise HTTPException(400, "Periode tidak valid (bulanan/sekali/semester)")
    if not (0 <= data.jatuh_tempo <= 31):
        raise HTTPException(400, "Jatuh tempo harus 1-31 (0 = tidak ada)")
    j = JenisPembayaran(**data.model_dump())
    db.add(j)
    db.commit()
    db.refresh(j)
    return j


@router.patch("/jenis/{jenis_id}", response_model=JenisPembayaranOut)
def update_jenis(
    jenis_id: int, data: JenisPembayaranUpdate,
    user: dict = Depends(require_permission("tagihan.kelola")),
    db: Session = Depends(get_tenant_db),
):
    """Edit jenis pembayaran."""
    j = db.get(JenisPembayaran, jenis_id)
    if not j:
        raise HTTPException(404, "Jenis pembayaran tidak ditemukan")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(j, k, v)
    db.commit()
    db.refresh(j)
    return j


@router.post("/jenis/{jenis_id}/toggle")
def toggle_jenis(
    jenis_id: int,
    user: dict = Depends(require_permission("tagihan.kelola")),
    db: Session = Depends(get_tenant_db),
):
    """Aktif/nonaktifkan jenis pembayaran."""
    j = db.get(JenisPembayaran, jenis_id)
    if not j:
        raise HTTPException(404, "Jenis pembayaran tidak ditemukan")
    j.is_active = not j.is_active
    db.commit()
    return {"id": j.id, "is_active": j.is_active}


# ─── Generate Tagihan ─────────────────────────────────────────────────────

@router.post("/generate")
def generate_tagihan(
    jenis_id: int | None = Query(None),
    periode: str = Query(..., description="YYYY-MM (bulan tagihan)"),
    user: dict = Depends(require_permission("tagihan.kelola")),
    db: Session = Depends(get_tenant_db),
):
    """Generate tagihan bulanan: semua murid aktif × jenis auto_generate.

    Idempotent: murid yang sudah punya tagihan di periode itu di-skip.
    Optional: filter jenis_id (kalau mau generate satu jenis saja).
    """
    try:
        datetime.strptime(periode, "%Y-%m")
    except ValueError:
        raise HTTPException(400, "Format periode harus YYYY-MM")

    q_jenis = db.query(JenisPembayaran).filter(JenisPembayaran.is_active == True)  # noqa: E712
    if jenis_id:
        q_jenis = q_jenis.filter(JenisPembayaran.id == jenis_id)
    jenis_list = q_jenis.all()

    total_baru = 0
    rincian = []
    for j in jenis_list:
        if j.periode != "bulanan" or not j.auto_generate:
            continue
        murids = (db.query(Murid)
                  .join(Kelas, Murid.kelas_id == Kelas.id)
                  .filter(Murid.is_active == True)  # noqa: E712
                  .all())
        baru = 0
        for m in murids:
            # Skip kalau sudah ada tagihan murid×jenis×periode
            ada = (db.query(Tagihan)
                   .filter_by(murid_id=m.id, jenis_id=j.id, periode=periode)
                   .first())
            if ada:
                continue
            # Jatuh tempo: tanggal j.jatuh_tempo pada bulan periode
            try:
                tempo = datetime(
                    int(periode[:4]), int(periode[5:7]), j.jatuh_tempo,
                    tzinfo=WIB)
            except ValueError:
                tempo = None
            db.add(Tagihan(murid_id=m.id, jenis_id=j.id, periode=periode,
                           nominal=j.nominal, jatuh_tempo=tempo,
                           status="belum"))
            baru += 1
        if baru:
            rincian.append(f"{j.nama}: {baru} tagihan")
        total_baru += baru
    db.commit()
    return {"total_baru": total_baru, "rincian": rincian}


# ─── Input Cepat (bulk) ───────────────────────────────────────────────────

@router.post("/bulk-bayar")
def bulk_bayar(
    entries: list[dict],
    user: dict = Depends(require_permission("tagihan.input")),
    db: Session = Depends(get_tenant_db),
):
    """Input pembayaran massal: [{"tagihan_id": 1, "nominal": 100000,
    "metode": "cash", "catatan": ""}, ...]. Satu transaksi, validasi semua dulu."""
    if not entries:
        raise HTTPException(400, "Tidak ada data pembayaran")
    guru_id = user.get("id") if user.get("role") == "guru" else user.get("guru_id")
    ok, gagal = 0, []
    for e in entries:
        t = db.get(Tagihan, e.get("tagihan_id"))
        if not t:
            gagal.append({"tagihan_id": e.get("tagihan_id"), "error": "Tagihan tidak ditemukan"})
            continue
        nominal = int(e.get("nominal", 0) or 0)
        if nominal <= 0:
            continue  # baris kosong → skip
        sisa = t.nominal - t.potongan - sum(p.nominal for p in t.pembayaran)
        if nominal > sisa:
            gagal.append({"tagihan_id": t.id, "error": f"Melebihi sisa (sisa Rp{sisa})"})
            continue
        if (not t.jenis.boleh_cicil and nominal < sisa):
            gagal.append({"tagihan_id": t.id, "error": "Jenis tidak boleh dicicil"})
            continue
        db.add(Pembayaran(tagihan_id=t.id, nominal=nominal,
                          metode=e.get("metode", "cash"),
                          guru_id=guru_id, catatan=e.get("catatan", "")))
        db.flush()
        _refresh_status(db, t)
        ok += 1
    db.commit()
    return {"ok": ok, "gagal": gagal}


@router.get("/export.xlsx")
def export_tagihan_xlsx(
    kelas_id: int | None = None,
    periode: str | None = None,
    status: str | None = None,
    user: dict = Depends(require_permission("tagihan.view", "tagihan.export", "tagihan.kelola")),
    db: Session = Depends(get_tenant_db),
):
    """Export rekap tagihan ke Excel (per murid per jenis per periode)."""
    import io

    from fastapi.responses import Response
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = _query_tagihan(db, kelas_id=kelas_id, periode=periode, status=status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Tagihan"

    # Identitas
    nama_aplikasi = "Aplikasi Madrasah"
    try:
        from ..models import Pengaturan
        p = db.query(Pengaturan).filter_by(key="nama_aplikasi").first()
        if p and p.value:
            nama_aplikasi = p.value
    except Exception:
        pass
    ws.append([nama_aplikasi])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"REKAP TAGIHAN — Periode: {periode or 'Semua'}"])
    ws.append([])

    headers = ["No", "NISN", "Nama", "Kelas", "Jenis", "Periode",
               "Nominal", "Potongan", "Terbayar", "Sisa", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=4, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F766E")
        c.alignment = Alignment(horizontal="center", vertical="center")

    from ..models import JenisPembayaran  # noqa: F401
    total_nominal = total_terbayar = 0
    for i, t in enumerate(rows, start=1):
        dibayar = sum(p.nominal for p in t.pembayaran)
        sisa = max(t.nominal - t.potongan - dibayar, 0)
        total_nominal += t.nominal
        total_terbayar += dibayar
        ws.append([
            i,
            t.murid.nisn if t.murid else "",
            t.murid.nama if t.murid else "",
            t.murid.kelas.nama_kelas if t.murid and t.murid.kelas else "",
            t.jenis.nama if t.jenis else "",
            t.periode,
            t.nominal, t.potongan, dibayar, sisa,
            {"belum": "Belum", "sebagian": "Sebagian", "lunas": "Lunas",
             "ditunda": "Ditunda"}.get(t.status, t.status),
        ])

    # Baris total
    ws.append([])
    ws.append(["", "", "", "", "", "TOTAL:", total_nominal, "",
               total_terbayar, max(total_nominal - total_terbayar, 0), ""])
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=ws.max_row, column=col)
        c.font = Font(bold=True)

    widths = [5, 14, 28, 8, 20, 10, 12, 10, 12, 12, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"rekap-tagihan-{periode or 'semua'}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# ─── List & Detail Tagihan ────────────────────────────────────────────────

def _query_tagihan(db: Session, kelas_id: int | None = None,
                   murid_id: int | None = None, periode: str | None = None,
                   status: str | None = None, jenis_id: int | None = None):
    q = (db.query(Tagihan)
         .options(joinedload(Tagihan.murid).joinedload(Murid.kelas),
                  joinedload(Tagihan.jenis),
                  joinedload(Tagihan.pembayaran).joinedload(Pembayaran.guru)))
    if kelas_id:
        q = (q.join(Murid, Tagihan.murid_id == Murid.id)
             .filter(Murid.kelas_id == kelas_id))
    if murid_id:
        q = q.filter(Tagihan.murid_id == murid_id)
    if periode:
        q = q.filter(Tagihan.periode == periode)
    if status:
        q = q.filter(Tagihan.status == status)
    if jenis_id:
        q = q.filter(Tagihan.jenis_id == jenis_id)
    return q.order_by(Tagihan.periode.desc(), Tagihan.murid_id).all()


@router.get("", response_model=list[TagihanOut])
def list_tagihan(
    kelas_id: int | None = None,
    murid_id: int | None = None,
    periode: str | None = None,
    status: str | None = None,
    jenis_id: int | None = None,
    user: dict = Depends(require_permission("tagihan.view")),
    db: Session = Depends(get_tenant_db),
):
    """Daftar tagihan (filter kelas/murid/periode/status/jenis)."""
    return [_tagihan_out(t) for t in _query_tagihan(
        db, kelas_id, murid_id, periode, status, jenis_id)]


@router.get("/rekap-kelas")
def rekap_kelas(
    kelas_id: int | None = None,
    periode: str | None = None,
    user: dict = Depends(require_permission("tagihan.view")),
    db: Session = Depends(get_tenant_db),
):
    """Rekap ringkas per kelas + per murid untuk periode tertentu."""
    rows = _query_tagihan(db, kelas_id=kelas_id, periode=periode)
    by_kelas: dict[int, dict] = {}
    for t in rows:
        k = t.murid.kelas if t.murid and t.murid.kelas else None
        k_id = k.id if k else 0
        k_nama = k.nama_kelas if k else "?"
        d = by_kelas.setdefault(k_id, {
            "kelas_id": k_id, "kelas": k_nama, "total": 0, "lunas": 0,
            "sebagian": 0, "belum": 0, "nominal": 0, "terbayar": 0})
        d["total"] += 1
        d["nominal"] += t.nominal
        dibayar = sum(p.nominal for p in t.pembayaran)
        d["terbayar"] += dibayar
        if t.status == "lunas":
            d["lunas"] += 1
        elif t.status == "sebagian":
            d["sebagian"] += 1
        else:
            d["belum"] += 1
    return sorted(by_kelas.values(), key=lambda x: x["kelas"])


@router.post("", response_model=TagihanOut, status_code=201)
def create_tagihan(
    data: TagihanCreate,
    user: dict = Depends(require_permission("tagihan.kelola", "tagihan.input")),
    db: Session = Depends(get_tenant_db),
):
    """Buat tagihan manual (untuk jenis sekali/semester atau kasus khusus)."""
    m = db.get(Murid, data.murid_id)
    j = db.get(JenisPembayaran, data.jenis_id)
    if not m:
        raise HTTPException(404, "Murid tidak ditemukan")
    if not j:
        raise HTTPException(404, "Jenis pembayaran tidak ditemukan")
    ada = (db.query(Tagihan)
           .filter_by(murid_id=m.id, jenis_id=j.id, periode=data.periode)
           .first())
    if ada:
        raise HTTPException(400, "Tagihan sudah ada untuk murid & periode ini")
    nominal = data.nominal if data.nominal else j.nominal
    t = Tagihan(murid_id=m.id, jenis_id=j.id, periode=data.periode,
                nominal=nominal, jatuh_tempo=data.jatuh_tempo, status="belum")
    db.add(t)
    db.commit()
    db.refresh(t)
    return _tagihan_out(t)


@router.get("/{tagihan_id}", response_model=TagihanOut)
def get_tagihan(
    tagihan_id: int,
    user: dict = Depends(require_permission("tagihan.view")),
    db: Session = Depends(get_tenant_db),
):
    """Detail tagihan + riwayat pembayaran."""
    t = db.get(Tagihan, tagihan_id)
    if not t:
        raise HTTPException(404, "Tagihan tidak ditemukan")
    _refresh_status(db, t)
    db.commit()
    db.refresh(t)
    return _tagihan_out(t)


# ─── Aksi: Bayar / Keringanan / Tunda ─────────────────────────────────────

@router.post("/{tagihan_id}/bayar", response_model=TagihanOut)
def bayar_tagihan(
    tagihan_id: int, data: TagihanBayarCreate,
    user: dict = Depends(require_permission("tagihan.input")),
    db: Session = Depends(get_tenant_db),
):
    """Input pembayaran: lunas (bayar penuh) atau cicil (bayar sebagian)."""
    t = db.get(Tagihan, tagihan_id)
    if not t:
        raise HTTPException(404, "Tagihan tidak ditemukan")
    if data.nominal <= 0:
        raise HTTPException(400, "Nominal harus lebih dari 0")
    sisa = t.nominal - t.potongan - sum(p.nominal for p in t.pembayaran)
    if data.nominal > sisa:
        raise HTTPException(400, f"Nominal melebihi sisa tagihan (sisa Rp{sisa})")

    # Cicilan tidak boleh kalau jenis tidak boleh dicicil & bukan lunas
    if (not t.jenis.boleh_cicil and data.nominal < sisa):
        raise HTTPException(400, "Jenis ini tidak boleh dicicil — harus bayar lunas")

    guru_id = user.get("id") if user.get("role") == "guru" else user.get("guru_id")
    p = Pembayaran(tagihan_id=t.id, nominal=data.nominal,
                   metode=data.metode, guru_id=guru_id, catatan=data.catatan)
    db.add(p)
    db.flush()          # autoflush=False → p harus di-flush dulu biar query lihat
    _refresh_status(db, t)
    db.commit()
    db.refresh(t)
    return _tagihan_out(t)


@router.post("/{tagihan_id}/keringanan", response_model=TagihanOut)
def keringanan_tagihan(
    tagihan_id: int, data: TagihanKeringananCreate,
    user: dict = Depends(require_permission("tagihan.input")),
    db: Session = Depends(get_tenant_db),
):
    """Keringanan: potongan Rp dari tagihan (kasus khusus per murid)."""
    t = db.get(Tagihan, tagihan_id)
    if not t:
        raise HTTPException(404, "Tagihan tidak ditemukan")
    if data.potongan < 0 or data.potongan >= t.nominal:
        raise HTTPException(400, "Potongan harus 0 < potongan < nominal")
    t.potongan = data.potongan
    t.catatan = data.catatan or t.catatan
    _refresh_status(db, t)
    db.commit()
    db.refresh(t)
    return _tagihan_out(t)


@router.post("/{tagihan_id}/tunda", response_model=TagihanOut)
def tunda_tagihan(
    tagihan_id: int, data: TagihanTundaCreate,
    user: dict = Depends(require_permission("tagihan.input")),
    db: Session = Depends(get_tenant_db),
):
    """Penundaan: geser jatuh tempo (status → ditunda sampai tanggal baru)."""
    t = db.get(Tagihan, tagihan_id)
    if not t:
        raise HTTPException(404, "Tagihan tidak ditemukan")
    t.ditunda_sampai = data.ditunda_sampai
    t.catatan = data.catatan or t.catatan
    _refresh_status(db, t)
    db.commit()
    db.refresh(t)
    return _tagihan_out(t)


@router.delete("/{tagihan_id}")
def delete_tagihan(
    tagihan_id: int,
    user: dict = Depends(require_permission("tagihan.kelola")),
    db: Session = Depends(get_tenant_db),
):
    """Hapus tagihan + pembayarannya (kasus salah generate)."""
    t = db.get(Tagihan, tagihan_id)
    if not t:
        raise HTTPException(404, "Tagihan tidak ditemukan")
    db.delete(t)
    db.commit()
    return {"ok": True}
