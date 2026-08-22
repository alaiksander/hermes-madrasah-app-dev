"""Router Jurnal Mengajar — API endpoint CRUD + absensi per-jam."""

from datetime import date, datetime, time
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import case, select, func
from sqlalchemy.orm import Session, joinedload

from ..deps import get_tenant_db, require_permission
from ..models import (
    Guru, GuruPengampu, Kelas, Murid, JurnalMengajar, JurnalAbsensi,
    TahunAjaran,
)
from ..schemas import (
    JurnalMengajarCreate, JurnalAbsensiUpdate, JurnalAbsensiBulkUpdate,
    JurnalMengajarOut, JurnalMengajarUpdate, JurnalListOut,
)

router = APIRouter(prefix="/api/jurnal", tags=["Jurnal Mengajar"])


def _guru_pengampu_kelas_ids(db: Session, guru_id: int) -> set[int]:
    """Kelas_id yang diampu guru tsb di TA aktif (untuk filter jurnal)."""
    ta_aktif = db.query(TahunAjaran).filter(
        TahunAjaran.is_active.is_(True)).first()
    if not ta_aktif:
        return set()
    rows = db.query(GuruPengampu.kelas_id).filter(
        GuruPengampu.guru_id == guru_id,
        GuruPengampuan.tahun_ajaran_id == ta_aktif.id,
        GuruPengampu.is_active.is_(True),
    ).all()
    return {k for (k,) in rows}


# ─── helpers ────────────────────────────────────────────────────────────────

def _jurnal_out(j: JurnalMengajar) -> dict:
    return {
        "id": j.id,
        "guru_id": j.guru_id,
        "guru_nama": j.guru.nama if j.guru else "",
        "kelas_nama": j.kelas.nama_kelas if j.kelas else "",
        "mata_pelajaran": j.mata_pelajaran,
        "tanggal": j.tanggal,
        "jam_mulai": j.jam_mulai,
        "jam_selesai": j.jam_selesai,
        "materi": j.materi,
        "catatan": j.catatan,
        "kelas_id": j.kelas_id,
        "status": j.status,
        "verified_by": j.verified_by,
        "verified_at": j.verified_at,
        "created_at": j.created_at,
        "updated_at": j.updated_at,
        "absensi": [
            {"id": a.id, "murid_id": a.murid_id, "status": a.status,
             "murid_nama": a.murid.nama if a.murid else ""}
            for a in j.absensi
        ],
    }


# ─── List ───────────────────────────────────────────────────────────────────

@router.get("", response_model=list[JurnalListOut])
def list_jurnal(
    dari: date | None = None,
    sampai: date | None = None,
    kelas_id: int | None = None,
    guru_id: int | None = None,
    status: str | None = None,
    user: dict = Depends(require_permission("jurnal.view")),
    db: Session = Depends(get_tenant_db),
):
    """List jurnal. Admin/BK: semua. Guru: hanya miliknya."""

    q = select(JurnalMengajar).options(
        joinedload(JurnalMengajar.guru),
        joinedload(JurnalMengajar.kelas),
    )

    # Akses: guru hanya lihat miliknya sendiri
    if user.get("role") == "guru":
        q = q.where(JurnalMengajar.guru_id == user["id"])
        # Filter by pengampu (guru cuma lihat jurnal untuk kelas yang dia ampu)
        pengampu_kelas = _guru_pengampu_kelas_ids(db, user["id"])
        if pengampu_kelas and kelas_id is None:
            q = q.where(JurnalMengajar.kelas_id.in_(pengampu_kelas))
        elif pengampu_kelas and kelas_id and kelas_id not in pengampu_kelas:
            q = q.where(False)  # kelas di luar scope
    elif guru_id:
        q = q.where(JurnalMengajar.guru_id == guru_id)

    if dari:
        q = q.where(JurnalMengajar.tanggal >= dari)
    if sampai:
        q = q.where(JurnalMengajar.tanggal <= sampai)
    if kelas_id:
        q = q.where(JurnalMengajar.kelas_id == kelas_id)
    if status:
        q = q.where(JurnalMengajar.status == status)

    q = q.order_by(JurnalMengajar.tanggal.desc(), JurnalMengajar.created_at.desc())

    rows = db.execute(q).scalars().unique().all()
    return [
        JurnalListOut(
            id=j.id,
            tanggal=j.tanggal,
            kelas_nama=j.kelas.nama_kelas if j.kelas else "",
            mata_pelajaran=j.mata_pelajaran,
            jam_mulai=j.jam_mulai,
            jam_selesai=j.jam_selesai,
            status=j.status,
            guru_nama=j.guru.nama if j.guru else "",
            created_at=j.created_at,
        )
        for j in rows
    ]


# ─── Create ─────────────────────────────────────────────────────────────────

@router.post("", response_model=JurnalMengajarOut, status_code=201)
def create_jurnal(
    data: JurnalMengajarCreate,
    user: dict = Depends(require_permission("jurnal.input")),
    db: Session = Depends(get_tenant_db),
):
    """Buat entri jurnal baru + auto-create JurnalAbsensi (semua murid hadir)."""

    # Validasi: kelas ada
    kelas = db.get(Kelas, data.kelas_id)
    if not kelas:
        raise HTTPException(400, "Kelas tidak ditemukan")

    # Validasi: guru_id dari token (guru only; admin/BK boleh set guru lain?)
    guru_id = user["id"]  # fallback
    if user.get("role") == "guru":
        guru_id = user["id"]

    # Validasi: guru hanya boleh create jurnal di kelas yang dia ampu
    if user.get("role") == "guru":
        pengampu_kelas = _guru_pengampu_kelas_ids(db, user["id"])
        if pengampu_kelas and data.kelas_id not in pengampu_kelas:
            raise HTTPException(403, "Anda tidak mengampu kelas ini — minta admin set penugasan")

    # Buat jurnal
    jurnal = JurnalMengajar(
        guru_id=guru_id,
        kelas_id=data.kelas_id,
        mata_pelajaran=data.mata_pelajaran,
        tanggal=data.tanggal,
        jam_mulai=data.jam_mulai,
        jam_selesai=data.jam_selesai,
        materi=data.materi,
        catatan=data.catatan,
        status="draft",
    )
    db.add(jurnal)
    db.flush()  # dapat id

    # Auto-create absensi: semua murid aktif di kelas itu = hadir
    murid_rows = db.execute(
        select(Murid).where(
            Murid.kelas_id == data.kelas_id,
            Murid.is_active == True,  # noqa: E712
        )
    ).scalars().all()

    for m in murid_rows:
        db.add(JurnalAbsensi(jurnal_id=jurnal.id, murid_id=m.id, status="hadir"))

    db.commit()
    db.refresh(jurnal)

    return _jurnal_out(jurnal)


# ─── Export ────────────────────────────────────────────────────────────────

def _jam_str(j: JurnalMengajar) -> str:
    if not j.jam_mulai:
        return ""
    s = j.jam_mulai.strftime("%H:%M")
    if j.jam_selesai:
        s += f"-{j.jam_selesai.strftime('%H:%M')}"
    return s


def _query_jurnal_export(db: Session, user: dict, dari: date | None,
                         sampai: date | None, kelas_id: int | None,
                         guru_id: int | None, status: str | None = None):
    """Query jurnal + absensi lengkap (dipakai export Excel & PDF)."""
    q = select(JurnalMengajar).options(
        joinedload(JurnalMengajar.guru),
        joinedload(JurnalMengajar.kelas),
        joinedload(JurnalMengajar.absensi).joinedload(JurnalAbsensi.murid),
    )
    if user.get("role") == "guru":
        q = q.where(JurnalMengajar.guru_id == user["id"])
    elif guru_id:
        q = q.where(JurnalMengajar.guru_id == guru_id)
    if dari:
        q = q.where(JurnalMengajar.tanggal >= dari)
    if sampai:
        q = q.where(JurnalMengajar.tanggal <= sampai)
    if kelas_id:
        q = q.where(JurnalMengajar.kelas_id == kelas_id)
    if status:
        q = q.where(JurnalMengajar.status == status)
    return db.execute(q).scalars().unique().all()


@router.get("/export.xlsx")
def export_jurnal_xlsx(
    dari: date | None = None,
    sampai: date | None = None,
    kelas_id: int | None = None,
    guru_id: int | None = None,
    status: str | None = None,
    rekap_absensi: bool = True,
    user: dict = Depends(require_permission("jurnal.export", "jurnal.view", "jurnal.verify")),
    db: Session = Depends(get_tenant_db),
):
    """Export jurnal mengajar ke Excel (per baris jurnal + rekap absensi opsional)."""
    import io

    from fastapi.responses import Response
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = _query_jurnal_export(db, user, dari, sampai, kelas_id, guru_id, status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jurnal Mengajar"

    # Header identitas
    ws.append(["LAPORAN JURNAL MENGAJAR"])
    ws["A1"].font = Font(bold=True, size=14)
    periode = f"{dari or '—'} s/d {sampai or '—'}" if (dari or sampai) else "Semua periode"
    ws.append([f"Periode: {periode}"])
    ws.append([])

    # Header tabel
    headers = ["No", "Tanggal", "Hari", "Kelas", "Mata Pelajaran",
               "Jam Mulai", "Jam Selesai", "Materi", "Catatan", "Status", "Guru"]
    if rekap_absensi:
        headers += ["Hadir", "Izin", "Sakit", "Alpa"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=4, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F766E")
        c.alignment = Alignment(horizontal="center", vertical="center")

    _HARI = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    _STATUS_LABEL = {"draft": "Draft", "submitted": "Disubmit", "verified": "Terverifikasi"}
    for i, j in enumerate(rows, start=1):
        row = [
            i, j.tanggal.strftime("%d/%m/%Y"),
            _HARI[j.tanggal.weekday()],
            j.kelas.nama_kelas if j.kelas else "",
            j.mata_pelajaran,
            j.jam_mulai.strftime("%H:%M") if j.jam_mulai else "",
            j.jam_selesai.strftime("%H:%M") if j.jam_selesai else "",
            j.materi, j.catatan or "",
            _STATUS_LABEL.get(j.status, j.status),
            j.guru.nama if j.guru else "",
        ]
        if rekap_absensi:
            rekap = {"hadir": 0, "izin": 0, "sakit": 0, "alpa": 0}
            for a in j.absensi:
                if a.status in rekap:
                    rekap[a.status] += 1
            row += [rekap["hadir"], rekap["izin"], rekap["sakit"], rekap["alpa"]]
        ws.append(row)

    # Lebar kolom
    widths = [5, 12, 8, 8, 22, 10, 10, 30, 25, 12, 18]
    if rekap_absensi:
        widths += [8, 8, 8, 8]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"jurnal-mengajar-{dari or 'semua'}-{sampai or 'periode'}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export.pdf")
def export_jurnal_pdf(
    dari: date | None = None,
    sampai: date | None = None,
    kelas_id: int | None = None,
    guru_id: int | None = None,
    status: str | None = None,
    rekap_absensi: bool = True,
    user: dict = Depends(require_permission("jurnal.export", "jurnal.view", "jurnal.verify")),
    db: Session = Depends(get_tenant_db),
):
    """Export jurnal mengajar ke PDF (header identitas + tabel + rekap absensi opsional)."""
    import io

    from fastapi.responses import Response
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    rows = _query_jurnal_export(db, user, dari, sampai, kelas_id, guru_id, status)

    # Nama aplikasi dari pengaturan tenant (sama seperti PDF absensi)
    nama_aplikasi = "Aplikasi Madrasah"
    try:
        from ..models import Pengaturan
        p = db.query(Pengaturan).filter_by(key="nama_aplikasi").first()
        if p and p.value:
            nama_aplikasi = p.value
    except Exception:
        pass

    styles = getSampleStyleSheet()
    s_judul = ParagraphStyle("judul", parent=styles["Title"], fontSize=14,
                             spaceAfter=2)
    s_sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9,
                           textColor=colors.grey, alignment=1, spaceAfter=2)
    s_kop = ParagraphStyle("kop", parent=styles["Normal"], fontSize=10,
                           alignment=1, spaceAfter=6)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=14 * mm, bottomMargin=12 * mm)
    story = []

    story.append(Paragraph(nama_aplikasi, s_judul))
    story.append(Paragraph("LAPORAN JURNAL MENGAJAR", s_sub))
    periode = f"{dari or '—'} s/d {sampai or '—'}" if (dari or sampai) else "Semua periode"
    story.append(Paragraph(f"Periode: {periode} &nbsp;·&nbsp; Jumlah entri: {len(rows)}", s_kop))
    story.append(Spacer(1, 4))

    _HARI = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    _STATUS_LABEL = {"draft": "Draft", "submitted": "Disubmit", "verified": "Terverifikasi"}
    data = [["No", "Tanggal", "Kelas", "Mapel", "Jam", "Materi", "Status"]]
    if rekap_absensi:
        data[0] += ["H", "I", "S", "A"]
    for i, j in enumerate(rows, start=1):
        row = [
            str(i), f"{j.tanggal.strftime('%d/%m')} {_HARI[j.tanggal.weekday()][:3]}",
            j.kelas.nama_kelas if j.kelas else "",
            j.mata_pelajaran, _jam_str(j), j.materi or "",
            _STATUS_LABEL.get(j.status, j.status),
        ]
        if rekap_absensi:
            rekap = {"hadir": 0, "izin": 0, "sakit": 0, "alpa": 0}
            for a in j.absensi:
                if a.status in rekap:
                    rekap[a.status] += 1
            row += [str(rekap["hadir"]), str(rekap["izin"]),
                    str(rekap["sakit"]), str(rekap["alpa"])]
        data.append(row)

    col_widths = [8 * mm, 24 * mm, 12 * mm, 28 * mm, 18 * mm, 60 * mm, 18 * mm]
    if rekap_absensi:
        col_widths += [8 * mm, 8 * mm, 8 * mm, 8 * mm]
    tbl = Table(data, repeatRows=1, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FDFA")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "H = Hadir · I = Izin · S = Sakit · A = Alpa (rekap absensi per sesi jurnal)",
        styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    filename = f"jurnal-mengajar-{dari or 'semua'}-{sampai or 'periode'}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Detail ─────────────────────────────────────────────────────────────────

@router.get("/{jurnal_id}", response_model=JurnalMengajarOut)
def get_jurnal(
    jurnal_id: int,
    user: dict = Depends(require_permission("jurnal.view")),
    db: Session = Depends(get_tenant_db),
):
    j = db.execute(
        select(JurnalMengajar)
        .options(
            joinedload(JurnalMengajar.guru),
            joinedload(JurnalMengajar.kelas),
            joinedload(JurnalMengajar.absensi).joinedload(JurnalAbsensi.murid),
        )
        .where(JurnalMengajar.id == jurnal_id)
    ).scalars().first()

    if not j:
        raise HTTPException(404, "Jurnal tidak ditemukan")

    # Cek akses: guru hanya bisa lihat miliknya
    if user.get("role") == "guru" and j.guru_id != user["id"]:
        raise HTTPException(403, "Bukan jurnal Anda")

    return _jurnal_out(j)


# ─── Update ─────────────────────────────────────────────────────────────────

@router.patch("/{jurnal_id}", response_model=JurnalMengajarOut)
def update_jurnal(
    jurnal_id: int,
    data: JurnalMengajarUpdate,
    user: dict = Depends(require_permission("jurnal.input")),
    db: Session = Depends(get_tenant_db),
):
    j = db.get(JurnalMengajar, jurnal_id)
    if not j:
        raise HTTPException(404, "Jurnal tidak ditemukan")

    if user.get("role") == "guru" and j.guru_id != user["id"]:
        raise HTTPException(403, "Bukan jurnal Anda")

    # Update only non-None fields
    for field in ("kelas_id", "mata_pelajaran", "tanggal",
                  "jam_mulai", "jam_selesai", "materi", "catatan"):
        val = getattr(data, field, None)
        if val is not None:
            setattr(j, field, val)
    j.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(j)
    return _jurnal_out(j)


# ─── Absensi Bulk Update ────────────────────────────────────────────────────

@router.post("/{jurnal_id}/absensi")
def update_absensi_bulk(
    jurnal_id: int,
    data: JurnalAbsensiBulkUpdate,
    user: dict = Depends(require_permission("jurnal.input")),
    db: Session = Depends(get_tenant_db),
):
    """Bulk update status absensi per-murid untuk satu jurnal."""
    j = db.get(JurnalMengajar, jurnal_id)
    if not j:
        raise HTTPException(404, "Jurnal tidak ditemukan")

    if user.get("role") == "guru" and j.guru_id != user["id"]:
        raise HTTPException(403, "Bukan jurnal Anda")

    valid = {"hadir", "izin", "sakit", "alpa"}
    updated = 0
    for murid_id_str, status_val in data.updates.items():
        if status_val not in valid:
            continue
        row = db.execute(
            select(JurnalAbsensi).where(
                JurnalAbsensi.jurnal_id == jurnal_id,
                JurnalAbsensi.murid_id == int(murid_id_str),
            )
        ).scalar_one_or_none()
        if row:
            row.status = status_val
            updated += 1

    db.commit()
    return {"updated": updated}


# ─── Submit (draft → submitted) ─────────────────────────────────────────────

@router.post("/{jurnal_id}/submit", response_model=JurnalMengajarOut)
def submit_jurnal(
    jurnal_id: int,
    user: dict = Depends(require_permission("jurnal.input")),
    db: Session = Depends(get_tenant_db),
):
    j = db.get(JurnalMengajar, jurnal_id)
    if not j:
        raise HTTPException(404, "Jurnal tidak ditemukan")

    if user.get("role") == "guru" and j.guru_id != user["id"]:
        raise HTTPException(403, "Bukan jurnal Anda")

    if j.status not in ("draft",):
        raise HTTPException(400, f"Status sudah {j.status}, tidak bisa disubmit")

    j.status = "submitted"
    j.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(j)
    return _jurnal_out(j)


# ─── Verify (submitted → verified) ─────────────────────────────────────────

@router.post("/{jurnal_id}/verify", response_model=JurnalMengajarOut)
def verify_jurnal(
    jurnal_id: int,
    user: dict = Depends(require_permission("jurnal.verify")),
    db: Session = Depends(get_tenant_db),
):
    j = db.get(JurnalMengajar, jurnal_id)
    if not j:
        raise HTTPException(404, "Jurnal tidak ditemukan")

    if j.status != "submitted":
        raise HTTPException(400, f"Status {j.status}, harus submitted dulu")

    j.status = "verified"
    j.verified_by = user["id"]
    j.verified_at = datetime.utcnow()
    db.commit()
    db.refresh(j)
    return _jurnal_out(j)


# ─── Dashboard stats ────────────────────────────────────────────────────────

@router.get("/stats/bulan-ini")
def stats_bulan_ini(
    user: dict = Depends(require_permission("jurnal.view")),
    db: Session = Depends(get_tenant_db),
):
    """Statistik jurnal bulan berjalan untuk dashboard.

    Return breakdown per status (draft / submitted / verified) supaya
    dashboard bisa tampilkan 4 cards konsisten dengan pola Absensi.
    """
    today = date.today()
    bulan = today.month
    tahun = today.year

    q = select(
        func.count(JurnalMengajar.id).label("total"),
        func.sum(
            case((JurnalMengajar.status == "draft", 1), else_=0)
        ).label("draft"),
        func.sum(
            case((JurnalMengajar.status == "submitted", 1), else_=0)
        ).label("submitted"),
        func.sum(
            case((JurnalMengajar.status == "verified", 1), else_=0)
        ).label("verified"),
    ).where(
        func.extract("month", JurnalMengajar.tanggal) == bulan,
        func.extract("year", JurnalMengajar.tanggal) == tahun,
    )

    if user.get("role") == "guru":
        q = q.where(JurnalMengajar.guru_id == user["id"])

    row = db.execute(q).one()
    total = row.total or 0
    return {
        "bulan": bulan,
        "tahun": tahun,
        "total_jurnal": total,
        "draft": row.draft or 0,
        "submitted": row.submitted or 0,
        "verified": row.verified or 0,
        "disubmit": (row.submitted or 0) + (row.verified or 0),  # backward-compat
    }
